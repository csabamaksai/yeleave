import re

with open('timesheet/views.py', 'r', encoding='utf-8') as f:
    content = f.read()

# Add get_filtered_report_entries helper
helper_code = '''
def get_filtered_report_entries(request):
    users = CustomUser.objects.filter(is_active=True).order_last_first()
    
    user_ids = request.GET.getlist('user_id')
    if user_ids:
        valid_user_ids = [int(u) for u in user_ids if u.isdigit()]
        selected_users = users.filter(id__in=valid_user_ids)
    else:
        selected_users = users.none()
        valid_user_ids = []

    try:
        year = int(request.GET.get('year', date.today().year))
        month = int(request.GET.get('month', date.today().month))
    except ValueError:
        year, month = date.today().year, date.today().month
        
    client_id = request.GET.get('client_id')
    project_id = request.GET.get('project_id')
    
    entries = []
    
    if selected_users.exists():
        qs = TimeEntry.objects.filter(
            user__in=selected_users,
            date__year=year,
            date__month=month
        ).select_related('project', 'project__client', 'user')
        
        if client_id:
            qs = qs.filter(project__client_id=client_id)
        if project_id:
            qs = qs.filter(project_id=project_id)
            
        for e in qs:
            entries.append({
                'is_leave': False,
                'user_name': f"{e.user.last_name} {e.user.first_name}",
                'date': e.date,
                'client_name': e.project.client.name,
                'project_name': e.project.name,
                'hours': e.hours,
                'description': e.description,
            })
            
        first_weekday, num_days = calendar.monthrange(year, month)
        start_of_month = date(year, month, 1)
        end_of_month = date(year, month, num_days)
        
        leaves = Leave.objects.filter(
            user__in=selected_users,
            start_date__lte=end_of_month,
            end_date__gte=start_of_month
        ).select_related('user')
        
        holidays = Holiday.objects.filter(
            date__year=year,
            date__month=month
        )
        holiday_dict = {h.date: h for h in holidays}
        
        for l in leaves:
            curr = l.start_date
            while curr <= l.end_date:
                if curr >= start_of_month and curr <= end_of_month:
                    is_calendar_weekend = curr.weekday() >= 5
                    is_weekend_behavior = is_calendar_weekend
                    if curr in holiday_dict:
                        holiday_obj = holiday_dict[curr]
                        if holiday_obj.is_working_day:
                            is_weekend_behavior = False
                        else:
                            is_weekend_behavior = True
                            
                    if not is_weekend_behavior:
                        entries.append({
                            'is_leave': True,
                            'user_name': f"{l.user.last_name} {l.user.first_name}",
                            'date': curr,
                            'client_name': '-',
                            'project_name': str(l.get_leave_type_display()),
                            'hours': 0,
                            'description': l.notes if l.notes else _("Szabadság"),
                        })
                curr += timedelta(days=1)
                
        entries.sort(key=lambda x: x['date'], reverse=True)

    return entries, year, month
'''

export_code = '''
import csv
import xlwt
import openpyxl
from openpyxl.utils import get_column_letter

def reports_export(request):
    if not (request.user.is_staff or getattr(request.user, 'is_reporter', False)):
        return HttpResponseForbidden(_("Nincs jogosultságod ehhez a felülethez."))

    entries, year, month = get_filtered_report_entries(request)
    
    # filter out leaves if checkbox not checked (default we include but if explicitly "include_leaves" is false)
    # Actually wait, in UI the leave filter is JS only. Let's add 'include_leaves' parameter.
    # We leave that for later. It defaults to include all.
    include_leaves = request.GET.get('include_leaves', 'true') == 'true'
    if not include_leaves:
        entries = [e for e in entries if not e['is_leave']]

    fmt = request.GET.get('format', 'csv')
    group_by_user = request.GET.get('group_by_user', 'false') == 'true'

    filename = f"yeleave_export_{year}_{month}"

    header = [_("Dolgozó"), _("Dátum"), _("Partner"), _("Projekt"), _("Óra"), _("Megjegyzés")]

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        response.write('\ufeff'.encode('utf8')) # BOM for Excel
        writer = csv.writer(response, delimiter=';')
        writer.writerow(header)
        for e in entries:
            writer.writerow([
                e['user_name'],
                e['date'].strftime('%Y-%m-%d'),
                e['client_name'],
                e['project_name'],
                e['hours'],
                e['description']
            ])
        return response

    elif fmt == 'xls':
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        
        if group_by_user:
            users_in_entries = set([e['user_name'] for e in entries])
            # limit sheet name length to 31
            for user in users_in_entries:
                ws = wb.add_sheet(user[:31])
                for col_num, h in enumerate(header):
                    ws.write(0, col_num, h)
                user_entries = [e for e in entries if e['user_name'] == user]
                for row_num, e in enumerate(user_entries, 1):
                    ws.write(row_num, 0, e['user_name'])
                    ws.write(row_num, 1, e['date'].strftime('%Y-%m-%d'))
                    ws.write(row_num, 2, e['client_name'])
                    ws.write(row_num, 3, e['project_name'])
                    ws.write(row_num, 4, e['hours'])
                    ws.write(row_num, 5, e['description'])
        else:
            ws = wb.add_sheet("Export")
            for col_num, h in enumerate(header):
                ws.write(0, col_num, h)
            for row_num, e in enumerate(entries, 1):
                ws.write(row_num, 0, e['user_name'])
                ws.write(row_num, 1, e['date'].strftime('%Y-%m-%d'))
                ws.write(row_num, 2, e['client_name'])
                ws.write(row_num, 3, e['project_name'])
                ws.write(row_num, 4, e['hours'])
                ws.write(row_num, 5, e['description'])
                
        wb.save(response)
        return response

    elif fmt == 'xlsx':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb = openpyxl.Workbook()
        
        if group_by_user and entries:
            wb.remove(wb.active) # Remove default sheet
            users_in_entries = set([e['user_name'] for e in entries])
            for user in users_in_entries:
                ws = wb.create_sheet(title=user[:31])
                ws.append(header)
                user_entries = [e for e in entries if e['user_name'] == user]
                for e in user_entries:
                    ws.append([
                        e['user_name'],
                        e['date'].strftime('%Y-%m-%d'),
                        e['client_name'],
                        e['project_name'],
                        e['hours'],
                        e['description']
                    ])
        else:
            ws = wb.active
            ws.title = "Export"
            ws.append(header)
            for e in entries:
                ws.append([
                    e['user_name'],
                    e['date'].strftime('%Y-%m-%d'),
                    e['client_name'],
                    e['project_name'],
                    e['hours'],
                    e['description']
                ])
                
        wb.save(response)
        return response
        
    return HttpResponse("Invalid format")
'''

content += f'\n\n{helper_code}\n\n{export_code}'

# Let's write the imports up top
content = re.sub(r'import calendar', 'import calendar\nimport csv\nimport xlwt\nimport openpyxl', content)
with open('timesheet/views.py', 'w', encoding='utf-8') as f:
    f.write(content)
