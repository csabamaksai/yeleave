from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, QueryDict, HttpResponseForbidden
from django.views.decorators.http import require_POST
import json
from calendar import monthrange
from datetime import date, datetime, timedelta
from itertools import groupby
from .models import TimeEntry
from projects.models import Project
from leaves.models import Leave
from holidays.models import Holiday
from users.models import CustomUser
import re

def format_hours(decimal_hours):
    if not decimal_hours:
        return ""
    h = int(decimal_hours)
    m = int(round((float(decimal_hours) - h) * 60))
    if m == 0:
        return f"{h}ó"
    elif h == 0:
        return f"{m}p"
    return f"{h}ó {m}p"

def parse_hours(val_str):
    val = val_str.strip().lower()
    if not val:
        return 0.0
    
    # Próbáljuk "7ó 40p" / "8ó" formátumból
    m_match = re.match(r'^(\d+)\s*(?:ó|o|h)\s*(?:(\d+)\s*(?:p|m)?)?$', val)
    if m_match:
        h = int(m_match.group(1))
        m_val = int(m_match.group(2)) if m_match.group(2) else 0
        return h + (m_val / 60.0)
    
    # Ha "p" vagy "m" re végződik, pl. 40p
    m_only = re.match(r'^(\d+)\s*(?:p|m)$', val)
    if m_only:
        return int(m_only.group(1)) / 60.0

    val = val.replace(',', '.')
    if ':' in val:
        parts = val.split(':')
        h = float(parts[0]) if parts[0] else 0.0
        m_val = float(parts[1]) if len(parts) > 1 and parts[1] else 0.0
        return h + (m_val / 60.0)

    try:
        return float(val)
    except ValueError:
        return 0.0

def get_prev_and_next_month(year, month):
    if month == 1:
        prev_month_date = date(year - 1, 12, 1)
        next_month_date = date(year, 2, 1)
    elif month == 12:
        prev_month_date = date(year, 11, 1)
        next_month_date = date(year + 1, 1, 1)
    else:
        prev_month_date = date(year, month - 1, 1)
        next_month_date = date(year, month + 1, 1)
    return prev_month_date, next_month_date

@login_required
def timesheet_view(request):
    try:
        year = int(request.GET.get('year', date.today().year))
        month = int(request.GET.get('month', date.today().month))
    except ValueError:
        year, month = date.today().year, date.today().month

    current_date = date(year, month, 1)
    prev_month, next_month = get_prev_and_next_month(year, month)

    # Naptári napok kiszámítása az adott hónapban
    _, num_days = monthrange(year, month)
    days_in_month = []
    
    # Keresünk szabadságot a felhasználónak, ami érinti ezt a hónapot
    leaves = Leave.objects.filter(
        user=request.user,
        start_date__lte=date(year, month, num_days),
        end_date__gte=date(year, month, 1)
    )
    # Keresünk ünnepeket erre a hónapra
    holidays = Holiday.objects.filter(
        date__year=year,
        date__month=month
    )
    holiday_dict = {h.date: h for h in holidays}
    
    # Létrehozunk egy set-et a gyors dátum kereséshez
    leave_dates = set()
    for l in leaves:
        curr = l.start_date
        while curr <= l.end_date:
            # Csak akkor számít szabadságnak vizuálisan a Timesheet-en, ha ez nem hétvége/ünnep, 
            # de ezt a logikát lejjebb oldjuk meg, vagy csak nem rakjuk be a leave_dates be.
            leave_dates.add(curr)
            curr += timedelta(days=1)
            
    for day in range(1, num_days + 1):
        day_date = date(year, month, day)
        date_str = str(day_date)
        
        # Alapértelmezetten hétvége-e
        is_calendar_weekend = day_date.weekday() >= 5
        
        is_weekend_behavior = is_calendar_weekend
        # Ha ünnepnap, akkor
        if day_date in holiday_dict:
            holiday_obj = holiday_dict[day_date]
            if holiday_obj.is_working_day:
                # Áthelyezett munkanap (pl szombat)
                is_weekend_behavior = False
            else:
                # Normál ünnepnap (hétköznap is lehet)
                is_weekend_behavior = True
        
        # Olyan nap ami igazi szabadság (nem esik hétvégére vagy ünnepnapra)
        is_leave_behavior = (day_date in leave_dates) and not is_weekend_behavior
        
        days_in_month.append({
            'date': day_date,
            'day': day,
            'is_weekend': is_weekend_behavior, 
            'weekday_name': day_date.strftime('%a')[:2], 
            'is_leave': is_leave_behavior 
        })

    # Felhasználóhoz rendelt projektek lekérdezése
    projects = request.user.assigned_projects.filter(is_active=True, client__is_active=True).select_related('client')
    
    # E havi mentett bejegyzések
    entries = TimeEntry.objects.filter(
        user=request.user,
        date__year=year,
        date__month=month
    )
    
    # Szótárat építünk a gyors kereséshez: (project_id, dátum_string) -> float hours
    entry_dict = {(e.project_id, str(e.date)): float(e.hours) for e in entries}

    # Sorok előkészítése a template-hez
    project_rows = []
    for project in projects:
        row_days = []
        for d in days_in_month:
            date_str = str(d['date'])
            hours_decimal = entry_dict.get((project.id, date_str), 0)
            hours_formatted = format_hours(hours_decimal) if hours_decimal > 0 else ''
            
            row_days.append({
                'date_str': date_str,
                'hours_formatted': hours_formatted,
                'hours_decimal': hours_decimal,
                'is_weekend': d['is_weekend'],
                'is_leave': d['is_leave']
            })
        project_rows.append({
            'project': project,
            'client_name': project.client.name,
            'days': row_days
        })

    return render(request, 'timesheet/index.html', {
        'current_date': current_date,
        'prev_month': prev_month,
        'next_month': next_month,
        'days_in_month': days_in_month,
        'project_rows': project_rows,
    })

@login_required
@require_POST
def save_time_entry(request):
    try:
        data = json.loads(request.body)
        project_id = data.get('project_id')
        entry_date_str = data.get('date')
        hours_str = data.get('hours')
        
        project = Project.objects.filter(id=project_id, assigned_users=request.user).first()
        if not project:
            return JsonResponse({'status': 'error', 'message': 'Project nem található vagy nincs rá jogosultságod.'}, status=403)

        entry_date = datetime.strptime(entry_date_str, '%Y-%m-%d').date()
        
        parsed_hours = parse_hours(hours_str) if hours_str else 0.0
        
        if parsed_hours <= 0:
            # Törölni kell, ha üres vagy 0 (mert nem rögzítünk 0 órát)
            TimeEntry.objects.filter(user=request.user, project=project, date=entry_date).delete()
            return JsonResponse({'status': 'success', 'message': 'Törölve', 'formatted_hours': ''})

        hours = round(parsed_hours, 2)
        
        # Frissítjük vagy Létrehozzuk
        entry, created = TimeEntry.objects.update_or_create(
            user=request.user,
            project=project,
            date=entry_date,
            defaults={'hours': hours}
        )
        
        return JsonResponse({'status': 'success', 'formatted_hours': format_hours(entry.hours)})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)



from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import get_object_or_404
from clients.models import Client
from leaves.models import Leave
from holidays.models import Holiday
import calendar
import csv
import xlwt
import openpyxl
from django.utils.translation import gettext as _

User = get_user_model()

def is_reporter_or_staff(user):
    return user.is_authenticated and (getattr(user, 'is_company_admin', False) or user.is_staff or user.is_superuser or getattr(user, 'is_reporter', False))

@user_passes_test(is_reporter_or_staff)
def reports_index(request):
    # Retrieve or save session state for reports
    if request.GET:
        # User explicitly requested a filter, save it
        request.session['reports_filters'] = request.GET.urlencode()
    else:
        # No filter provided (e.g. clicked on menu), try loading from session
        saved_filters = request.session.get('reports_filters')
        if saved_filters:
            request.GET = QueryDict(saved_filters)

    users = list(User.objects.filter(is_active=True, is_staff=False, is_superuser=False).order_by('last_name', 'first_name').prefetch_related('assigned_projects', 'assigned_projects__client'))
    for u in users:
        active_projs = [p for p in u.assigned_projects.all() if p.is_active and p.client.is_active]
        u.data_project_ids = ','.join(str(p.id) for p in active_projs)
        u.data_client_ids = ','.join(str(p.client_id) for p in active_projs)
    
    user_ids = request.GET.getlist('user_id')
    if user_ids:
        valid_user_ids = [int(u) for u in user_ids if u.isdigit()]
        selected_users = [u for u in users if u.id in valid_user_ids]
    else:
        selected_users = []
        valid_user_ids = []
        
    user_query_params = "".join([f"&user_id={uid}" for uid in valid_user_ids])

    try:
        year = int(request.GET.get('year', date.today().year))
        month = int(request.GET.get('month', date.today().month))
    except ValueError:
        year, month = date.today().year, date.today().month
        
    prev_month_date, next_month_date = get_prev_and_next_month(year, month)
    
    entries, _, _ = get_filtered_report_entries(request)
    
    client_id = request.GET.get('client_id')
    project_id = request.GET.get('project_id')
    
    clients = Client.objects.filter(is_active=True).order_by('name')
    projects = Project.objects.filter(is_active=True, client__is_active=True).order_by('name')
    # Removed the projects filtering here so the template gets all projects for the JS filter
        
    total_hours = sum(float(e['hours']) for e in entries if not e['is_leave'])
    
    # Sort and group entries by user for the template
    entries.sort(key=lambda x: (x['user_name'], -x['date'].toordinal()))
    grouped_entries = []
    
    for user_name, group in groupby(entries, key=lambda x: x['user_name']):
        group_list = list(group)
        user_hours = sum(float(e['hours']) for e in group_list if not e['is_leave'])
        user_days = round(user_hours / 8, 2)
        grouped_entries.append({
            'user_name': user_name,
            'entries': group_list,
            'total_hours': user_hours,
            'total_days': user_days
        })
    
    active_project_ids = []
    active_client_ids = []
    if selected_users:
        active_project_ids = list(TimeEntry.objects.filter(user__in=[su.id for su in selected_users]).values_list('project_id', flat=True).distinct())
        active_client_ids = list(Project.objects.filter(id__in=active_project_ids).values_list('client_id', flat=True).distinct())

    # Get users that are assigned to the currently selected project/client
    # If project selected -> users assigned to this project
    # If only client selected -> users assigned to ANY project belonging to this client
    # If neither -> all users are "assigned_user_ids" conceptually (or we just don't fade any)
    assigned_user_ids = []
    if project_id:
        p = Project.objects.filter(id=project_id).first()
        if p:
            assigned_user_ids = list(p.assigned_users.values_list('id', flat=True))
    elif client_id:
        assigned_user_ids = list(User.objects.filter(assigned_projects__client_id=client_id, is_staff=False, is_superuser=False).values_list('id', flat=True).distinct())
    else:
        assigned_user_ids = [u.id for u in users]

    context = {
        'grouped_entries': grouped_entries,
        'assigned_user_ids': assigned_user_ids,
        'users': users,
        'valid_user_ids': valid_user_ids,
        'user_query_params': user_query_params,
        'entries': entries,
        'total_hours': total_hours,
        'total_days': total_hours / 8 if total_hours else 0,
        'year': year,
        'month': month,
        'prev_month': prev_month_date,
        'next_month': next_month_date,
        'clients': clients,
        'projects': projects,
        'selected_client_id': client_id,
        'selected_project_id': project_id,
        'current_date': date(year, month, 1),
        'active_project_ids': active_project_ids,
        'active_client_ids': active_client_ids,
    }
    return render(request, 'reports/index.html', context)



def get_filtered_report_entries(request):
    users = CustomUser.objects.filter(is_active=True, is_staff=False, is_superuser=False).order_by('last_name', 'first_name')
    
    user_ids = request.GET.getlist('user_id')
    if user_ids:
        valid_user_ids = [int(u) for u in user_ids if u.isdigit()]
        selected_users = [u for u in users if u.id in valid_user_ids]
    else:
        selected_users = []
        valid_user_ids = []

    try:
        year = int(request.GET.get('year', date.today().year))
        month = int(request.GET.get('month', date.today().month))
    except ValueError:
        year, month = date.today().year, date.today().month
        
    client_id = request.GET.get('client_id')
    project_id = request.GET.get('project_id')
    
    entries = []
    
    if selected_users:
        qs = TimeEntry.objects.filter(
            user__in=selected_users,
            date__year=year,
            date__month=month
        ).select_related('project', 'project__client', 'user')
        
        if client_id:
            qs = qs.filter(project__client_id=client_id)
        if project_id:
            qs = qs.filter(project_id=project_id)
            
        for e in qs:
            entries.append({
                'is_leave': False,
                'user_name': f"{e.user.last_name} {e.user.first_name}",
                'date': e.date,
                'client_name': e.project.client.name,
                'project_name': e.project.name,
                'hours': e.hours,
                'description': e.description,
            })
            
        first_weekday, num_days = calendar.monthrange(year, month)
        start_of_month = date(year, month, 1)
        end_of_month = date(year, month, num_days)
        
        leaves = Leave.objects.filter(
            user__in=selected_users,
            start_date__lte=end_of_month,
            end_date__gte=start_of_month
        ).select_related('user')
        
        holidays = Holiday.objects.filter(
            date__year=year,
            date__month=month
        )
        holiday_dict = {h.date: h for h in holidays}
        
        for l in leaves:
            curr = l.start_date
            while curr <= l.end_date:
                if curr >= start_of_month and curr <= end_of_month:
                    is_calendar_weekend = curr.weekday() >= 5
                    is_weekend_behavior = is_calendar_weekend
                    if curr in holiday_dict:
                        holiday_obj = holiday_dict[curr]
                        if holiday_obj.is_working_day:
                            is_weekend_behavior = False
                        else:
                            is_weekend_behavior = True
                            
                    if not is_weekend_behavior:
                        entries.append({
                            'is_leave': True,
                            'user_name': f"{l.user.last_name} {l.user.first_name}",
                            'date': curr,
                            'client_name': '-',
                            'project_name': str(l.get_leave_type_display()),
                            'hours': 0,
                            'description': l.notes if l.notes else _("Szabadság"),
                        })
                curr += timedelta(days=1)
                
        entries.sort(key=lambda x: x['date'], reverse=True)

    return entries, year, month



import csv
import xlwt
import openpyxl
from openpyxl.utils import get_column_letter

def reports_export(request):
    if not (getattr(request.user, 'is_company_admin', False) or request.user.is_staff or request.user.is_superuser or getattr(request.user, 'is_reporter', False)):
        return HttpResponseForbidden(_("Nincs jogosultságod ehhez a felülethez."))

    entries, year, month = get_filtered_report_entries(request)
    
    # filter out leaves if checkbox not checked (default we include but if explicitly "include_leaves" is false)
    # Actually wait, in UI the leave filter is JS only. Let's add 'include_leaves' parameter.
    # We leave that for later. It defaults to include all.
    include_leaves = request.GET.get('include_leaves', 'true') == 'true'
    if not include_leaves:
        entries = [e for e in entries if not e['is_leave']]

    fmt = request.GET.get('format', 'csv')
    group_by_user = request.GET.get('group_by_user', 'false') == 'true'

    filename = f"yeleave_export_{year}_{month}"

    header = [_("Tanácsadó"), _("Dátum"), _("Partner"), _("Projekt"), _("Óra"), _("Megjegyzés")]

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        response.write('﻿'.encode('utf8')) # BOM for Excel
        writer = csv.writer(response, delimiter=';')
        writer.writerow(header)
        for e in entries:
            writer.writerow([
                e['user_name'],
                e['date'].strftime('%Y-%m-%d'),
                e['client_name'],
                e['project_name'],
                e['hours'],
                e['description']
            ])
        return response

    elif fmt == 'xls':
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        
        if group_by_user:
            users_in_entries = set([e['user_name'] for e in entries])
            # limit sheet name length to 31
            for user in users_in_entries:
                ws = wb.add_sheet(user[:31])
                for col_num, h in enumerate(header):
                    ws.write(0, col_num, h)
                user_entries = [e for e in entries if e['user_name'] == user]
                for row_num, e in enumerate(user_entries, 1):
                    ws.write(row_num, 0, e['user_name'])
                    ws.write(row_num, 1, e['date'].strftime('%Y-%m-%d'))
                    ws.write(row_num, 2, e['client_name'])
                    ws.write(row_num, 3, e['project_name'])
                    ws.write(row_num, 4, e['hours'])
                    ws.write(row_num, 5, e['description'])
        else:
            ws = wb.add_sheet("Export")
            for col_num, h in enumerate(header):
                ws.write(0, col_num, h)
            for row_num, e in enumerate(entries, 1):
                ws.write(row_num, 0, e['user_name'])
                ws.write(row_num, 1, e['date'].strftime('%Y-%m-%d'))
                ws.write(row_num, 2, e['client_name'])
                ws.write(row_num, 3, e['project_name'])
                ws.write(row_num, 4, e['hours'])
                ws.write(row_num, 5, e['description'])
                
        wb.save(response)
        return response

    elif fmt == 'xlsx':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb = openpyxl.Workbook()
        
        if group_by_user and entries:
            wb.remove(wb.active) # Remove default sheet
            users_in_entries = set([e['user_name'] for e in entries])
            for user in users_in_entries:
                ws = wb.create_sheet(title=user[:31])
                ws.append(header)
                user_entries = [e for e in entries if e['user_name'] == user]
                for e in user_entries:
                    ws.append([
                        e['user_name'],
                        e['date'].strftime('%Y-%m-%d'),
                        e['client_name'],
                        e['project_name'],
                        e['hours'],
                        e['description']
                    ])
        else:
            ws = wb.active
            ws.title = "Export"
            ws.append(header)
            for e in entries:
                ws.append([
                    e['user_name'],
                    e['date'].strftime('%Y-%m-%d'),
                    e['client_name'],
                    e['project_name'],
                    e['hours'],
                    e['description']
                ])
                
        wb.save(response)
        return response
        
    return HttpResponse("Invalid format")

@login_required
@require_POST
def bulk_save_time_entries(request):
    try:
        data = json.loads(request.body)
        entries = data.get('entries', [])
        
        # Lehetőségek: validálni a projekt jogosultságot bulkosan
        updated_count = 0
        deleted_count = 0
        for entry in entries:
            project_id = entry.get('project_id')
            entry_date_str = entry.get('date')
            hours_str = entry.get('hours')
            
            project = Project.objects.filter(id=project_id, assigned_users=request.user).first()
            if not project:
                continue
                
            entry_date = datetime.strptime(entry_date_str, '%Y-%m-%d').date()
            parsed_hours = parse_hours(hours_str) if hours_str else 0.0
            
            if parsed_hours <= 0:
                # Törölni kell, ha 0
                deleted = TimeEntry.objects.filter(user=request.user, project=project, date=entry_date).delete()
                if deleted[0] > 0:
                    deleted_count += deleted[0]
            else:
                hours = round(parsed_hours, 2)
                obj, created = TimeEntry.objects.update_or_create(
                    user=request.user,
                    project=project,
                    date=entry_date,
                    defaults={'hours': hours}
                )
                updated_count += 1
                
        return JsonResponse({'status': 'success', 'message': f'Sikeresen beküldve: {updated_count} módosítva/létrehozva, {deleted_count} törölve.'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

from django.contrib.auth.decorators import login_required, user_passes_test
import json
from django.http import JsonResponse
from datetime import datetime, date
from timesheet.models import ClientCertificate
from clients.models import Client

def is_company_admin_only(user):
    return getattr(user, 'is_company_admin', False) or user.is_staff or user.is_superuser

@login_required
@user_passes_test(is_company_admin_only)
def partner_tig_index(request):
    # Retrieve or save session state for partner tig
    if request.GET:
        request.session['partner_tig_filters'] = request.GET.urlencode()
    else:
        saved_filters = request.session.get('partner_tig_filters')
        if saved_filters:
            request.GET = QueryDict(saved_filters)

    try:
        year = int(request.GET.get('year', datetime.now().year))
        month = int(request.GET.get('month', datetime.now().month))
    except ValueError:
        year = datetime.now().year
        month = datetime.now().month

    client_id = request.GET.get('client_id')
    
    clients = Client.objects.filter(is_active=True).order_by('name')
    
    users = []
    if client_id:
        # Get existing certificates for the month and client
        certificates = ClientCertificate.objects.filter(
            year=year,
            month=month,
            client_id=client_id
        ).select_related('user')
        
        cert_dict = {c.user_id: c for c in certificates}
        
        # Get users who have time entries for this client in this month
        time_entries = TimeEntry.objects.filter(
            date__year=year, 
            date__month=month,
            project__client_id=client_id
        ).select_related('user')
        
        user_dict = {}
        # First, add everyone who has a certificate
        for c in certificates:
            user_dict[c.user_id] = {
                'user': c.user,
                'internal_hours': 0
            }
            
        # Then compile time entries (and add new users if they have time entries but no cert yet)
        for te in time_entries:
            if te.user_id not in user_dict:
                user_dict[te.user_id] = {
                    'user': te.user,
                    'internal_hours': 0
                }
            user_dict[te.user_id]['internal_hours'] += te.hours

        
        # Sort the dictionary keys by user's full name to display alphabetically
        sorted_users = sorted(user_dict.keys(), key=lambda uid: (user_dict[uid]['user'].last_name, user_dict[uid]['user'].first_name))
        
        for u_id in sorted_users:
            data = user_dict[u_id]
            cert = cert_dict.get(u_id)
            users.append({
                'user': data['user'],
                'internal_hours': float(data['internal_hours']),
                'internal_days': round(float(data['internal_hours']) / 8, 2),
                'cert_value': float(cert.value) if cert else '',
                'cert_unit': cert.unit if cert else 'days',
                'cert_notes': cert.notes if cert else ''
            })
            
    # prev/next month
    if month == 1:
        prev_month = {'year': year - 1, 'month': 12}
    else:
        prev_month = {'year': year, 'month': month - 1}

    if month == 12:
        next_month = {'year': year + 1, 'month': 1}
    else:
        next_month = {'year': year, 'month': month + 1}
        
    selected_client_name = ""
    if client_id and client_id.isdigit():
        for c in clients:
            if c.id == int(client_id):
                selected_client_name = c.name
                break

    context = {
        'year': year,
        'month': month,
        'current_date': date(year, month, 1),
        'clients': clients,
        'selected_client_id': int(client_id) if client_id and client_id.isdigit() else '',
        'selected_client_name': selected_client_name,
        'users': users,
        'prev_month': prev_month,
        'next_month': next_month,
    }
    return render(request, 'timesheet/partner_tig.html', context)

@login_required
@user_passes_test(is_company_admin_only)
def api_partner_tig_save(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            year = data.get('year')
            month = data.get('month')
            client_id = data.get('client_id')
            entries = data.get('entries', [])
            
            if not all([year, month, client_id]):
                return JsonResponse({'status': 'error', 'message': 'Missing required fields'})
                
            for entry in entries:
                user_id = entry.get('user_id')
                value = entry.get('value')
                unit = entry.get('unit', 'days')
                notes = entry.get('notes', '')
                
                if value == '' or value is None:
                    # If empty, delete the cert if exists
                    ClientCertificate.objects.filter(year=year, month=month, client_id=client_id, user_id=user_id).delete()
                else:
                    ClientCertificate.objects.update_or_create(
                        year=year,
                        month=month,
                        client_id=client_id,
                        user_id=user_id,
                        defaults={
                            'value': value,
                            'unit': unit,
                            'notes': notes
                        }
                    )
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
